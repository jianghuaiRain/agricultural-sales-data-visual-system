
import datetime  # 时间戳/格式化
import os        # 文件路径/存在性判断
import random    # 随机数（延迟/行为模拟）
import re        # 正则表达式（提取URL）
import time      # 延迟等待
import requests  # HTTP请求
from lxml import etree  # HTML解析（xpath）
import openpyxl  # Excel读写


# 定义请求头（包含Cookie，模拟登录状态）
headers = {
    'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/142.0.0.0 Safari/537.36 Edg/142.0.0.0',
    'Cookie': 'deviceId=e2e918e-5c55-48db-8400-c171803cf; Hm_lvt_b99541cbfb0edd202bb49abf3a0bef84=1762239771,1762417484; Hm_lvt_b2daa4a53a78af99c4a57c440f46d069=1763467426; sessionId=S_0MJ14V2VY4TLN49D; Hm_lvt_0e023fed85d2150e7d419b5b1f2e7c0f=1763467071,1765439109; HMACCOUNT=F1D67AF373D4EF14; Hm_lvt_91cf34f62b9bedb16460ca36cf192f4c=1763463548,1765439110; Hm_lvt_a6458082fb548e5ca7ff77d177d2d88d=1763463548,1765439110; hnUserTicket=58ed88d8-a4de-4a78-9abd-d52fd8eaf898; hnUserId=781226109; tfstk=gbnnpXxZr2zBMWSC-LqB8lRI2NLOAkZ77bI8wuFy75P196IKU70oOfp7zBZKqfcs1XHKekWI5jHVvDLQ2gqQVuRvMnKxpvZ74hW7n55C78w22MPFYkZZ7iliMnKxdvWLLKHwD2LNzdyzauzUzNSa1-Pz477eI5yuha5EauJMI-wR4azUTN-aU8EzabryIAP_UuyrauJiQ5wrMZ_UO0oKbQUmdP7q6nmaKyVqLWkKBcW38wMasgjoj2z3ggVG4gogKYzTXhSDySub9rFs_hI3qAyoi-ch_CqrIDi07DR1qjciul2s5Qbui4kKl4zMLHkgxW4mXyxW9k0rO04I-T9jQklslSae5CwivmUuGPfwTAHgTr0o9Csz9qDZsrnO6iEZ3VqG4r7NuFnfVRJ-ja_78RwgMhMWwj6C6KOJIdbVFyy_LSpMIa_78RwgMdvGogaUCJPA.; Hm_lpvt_a6458082fb548e5ca7ff77d177d2d88d=1765439168; Hm_lpvt_91cf34f62b9bedb16460ca36cf192f4c=1765439176; Hm_lpvt_0e023fed85d2150e7d419b5b1f2e7c0f=1765439176',
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8',
    'Accept-Language': 'zh-CN,zh;q=0.9,en;q=0.8',
    'Referer': 'https://www.cnhnb.com/',
    'Connection': 'keep-alive'
}

# 定义所有品类的基础URL
category_urls = {
    "蔬菜": "https://www.cnhnb.com/p/sczw/",
    "禽畜肉蛋": "https://www.cnhnb.com/p/qcrd/",
    "农副加工": "https://www.cnhnb.com/p/nfjg/",
    "水产": "https://www.cnhnb.com/p/shuic/",
    "粮油米面": "https://www.cnhnb.com/p/lymm/",
    "水果": "https://www.cnhnb.com/p/sgzw/",
    "种子种苗": "https://www.cnhnb.com/p/zzzm/",
}


def print_real_time_data(category, title, url, index):
    """实时打印爬取到的数据"""
    current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    print(f"\n{'=' * 80}")
    print(f"【爬取时间】: {current_time}")
    print(f"【品类】    : {category}")
    print(f"【序号】    : {index}")
    print(f"【标题】    : {title}")
    print(f"【链接】    : {url}")
    print(f"{'=' * 80}\n")


def check_existing_data(filename='惠农网全品类数据.xlsx'):
    """检查Excel中已有的URL，避免重复"""
    existing_urls = set()
    if not os.path.exists(filename):
        return existing_urls

    try:
        wb = openpyxl.load_workbook(filename)
        ws = wb.active
        # 读取第3列（URL列）的所有数据
        for row in range(2, ws.max_row + 1):
            url_cell = ws.cell(row=row, column=3).value
            if url_cell:
                existing_urls.add(url_cell)
        wb.close()
    except Exception as e:
        print(f"读取现有数据失败：{e}")

    return existing_urls


def write_to_excel(data, filename='惠农网全品类数据.xlsx', is_append=True, category=None, force_add=False):
    """写入Excel文件，支持增量写入和强制新增"""
    try:
        # 检测文件是否被占用
        if os.path.exists(filename):
            try:
                with open(filename, 'a') as f:
                    pass
            except PermissionError:
                print(f'错误：文件 {filename} 被占用（可能Excel已打开），请关闭后重试！')
                return 0

        if os.path.exists(filename) and is_append:
            wb = openpyxl.load_workbook(filename)
            ws = wb.active
            existing_urls = set()

            # 非强制新增时，读取现有URL去重
            if not force_add:
                for row in range(2, ws.max_row + 1):
                    url_cell = ws.cell(row=row, column=3).value
                    if url_cell:
                        existing_urls.add(url_cell)

            new_data_count = 0
            current_total = ws.max_row - 1

            for item in data:
                category_name, title, url = item
                if force_add or url not in existing_urls:
                    current_total += 1
                    ws.append([category_name, title, url])
                    new_data_count += 1
                    print(f"【新增数据】- 累计第 {current_total} 条")
                    print(f"品类: {category_name}")
                    print(f"标题: {title}")
                    print(f"链接: {url}")
                    print("-" * 60)

            wb.save(filename)
            wb.close()
            print(f'写入完成：新增 {new_data_count} 条数据，总数据量: {current_total} 条')
            return new_data_count
        else:
            # 新建文件
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(['品类', '标题', '链接'])

            for i, item in enumerate(data, 1):
                category_name, title, url = item
                ws.append([category_name, title, url])
                print(f"【新增数据】- 累计第 {i} 条")
                print(f"品类: {category_name}")
                print(f"标题: {title}")
                print(f"链接: {url}")
                print("-" * 60)

            wb.save(filename)
            wb.close()
            print(f'新建文件并写入：共 {len(data)} 条数据')
            return len(data)
    except Exception as e:
        print(f'写入Excel文件时出错: {e}')
        return 0


def scrape_category(category, base_url):
    """爬取单个品类的分页数据（修复分页URL构建逻辑）"""
    all_data = []
    category_total = 0
    # 读取现有URL
    existing_urls = check_existing_data('惠农网全品类数据.xlsx')

    # 提取品类标识（如sczw/qcrd，用于构建正确分页URL）
    category_code = base_url.split('/')[-2]

    # 爬取
    for i in range(101, 151):
        try:
            # 偶尔打乱节奏：每爬5页随机访问首页
            if i % 5 == 0 and random.random() > 0.5:
                print(f'模拟用户行为：随机访问首页...')
                requests.get('https://www.cnhnb.com/', headers=headers, timeout=10)
                time.sleep(random.randint(3, 5))

            # 构建正确的分页URL（核心修复！匹配网站真实规则：品类标识-0-0-0-0-页码）
            page_url = f'https://www.cnhnb.com/p/{category_code}-0-0-0-0-{i}/'

            print(f'正在爬取【{category}】第{i}页 - URL: {page_url}')
            response = requests.get(page_url, headers=headers, timeout=10)
            response.encoding = response.apparent_encoding  # 自动识别编码

            # 验证页面是否有效（可选）
            if response.status_code != 200:
                print(f' 【{category}】第{i}页访问失败，状态码：{response.status_code}')
                continue

            html_content = response.text
            # 提取商品ID和标题
            urls_num = re.findall(r'href="/gongying/(\d+)/"', html_content)
            title_elements = etree.HTML(html_content).xpath('//div[@class="shop-image"]//img/@title')
            titles = [title.strip() for title in title_elements if title.strip()]
            urls = [f'https://www.cnhnb.com/gongying/{num}/' for num in urls_num]
            min_length = min(len(titles), len(urls))

            # 统计新数据量
            new_count = 0
            for j in range(min_length):
                title = titles[j]
                url = urls[j]
                if url not in existing_urls:
                    new_count += 1
                    all_data.append((category, title, url))
                    category_total += 1
                    # 实时打印新数据
                    print_real_time_data(category, title, url, category_total)
                else:
                    print(f'重复数据：{title} | {url}')

            print(f'【{category}】第{i}页完成：{min_length} 条数据，新数据 {new_count} 条，累计: {category_total} 条')

            # 随机延迟（3-8秒）
            sleep_time = random.randint(3, 8)
            print(f'延迟 {sleep_time} 秒后继续...')
            time.sleep(sleep_time)

        except Exception as e:
            print(f'【{category}】第{i}页爬取失败: {str(e)}')
            # 重试一次
            time.sleep(5)
            try:
                response = requests.get(page_url, headers=headers, timeout=10)
                print(f'重试成功！')
            except:
                print(f'重试失败，跳过该页')
                break

    return all_data


def scrape_all_categories():
    """爬取所有品类的主函数"""
    total_saved = 0
    filename = '惠农网全品类数据.xlsx'
    total_categories = len(category_urls)
    current_category_index = 1

    print("=" * 80)
    print("惠农网全品类数据爬取程序启动")
    print(f"启动时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"待爬取品类总数：{total_categories} 个")
    print("=" * 80 + "\n")

    for category, base_url in category_urls.items():
        print(f"\n{'=' * 80}")
        print(f"开始爬取品类 ({current_category_index}/{total_categories})：{category}")
        print(f"基础URL：{base_url}")
        print("=" * 80)

        category_data = scrape_category(category, base_url)
        if category_data:
            # 可选：force_add=True 强制新增（忽略重复）
            # saved_count = write_to_excel(category_data, filename, category=category, force_add=False)
            saved_count = write_to_excel(category_data, filename, is_append=True, category=category, force_add=False)
            total_saved += saved_count
            print(f'\n【{category}】爬取完成！本次保存 {saved_count} 条数据')
            print(f'累计保存总数据量：{total_saved} 条\n')
        else:
            print(f'\n【{category}】未爬取到有效数据\n')

        current_category_index += 1

    print("=" * 80)
    print("所有品类爬取完成！")
    print(f"结束时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"最终累计保存数据总量：{total_saved} 条")
    print(f"数据文件：{os.path.abspath(filename)}")
    print("=" * 80)


if __name__ == "__main__":
    scrape_all_categories()