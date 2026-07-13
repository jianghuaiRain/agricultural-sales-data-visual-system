import requests
from lxml import etree
import random
import re
import time
import openpyxl
import os

# 定义URL和请求头
base_url = 'https://www.cnhnb.com/p/sczw/'  # 替换为实际链接1
headers = {
    'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/142.0.0.0 Safari/537.36 Edg/142.0.0.0'}


def write_to_excel(data, filename='惠农网数据.xlsx', is_append=True):
    """将数据写入Excel文件，支持增量写入"""
    try:
        if os.path.exists(filename) and is_append:
            # 增量写入模式
            wb = openpyxl.load_workbook(filename)
            ws = wb.active

            # 获取已存在的URL集合（用于去重）
            existing_urls = set()
            for row in range(2, ws.max_row + 1):  # 从第2行开始（跳过表头）
                url_cell = ws.cell(row=row, column=2).value  # 第2列是URL
                if url_cell:
                    existing_urls.add(url_cell)

            # 只添加新的数据
            new_data_count = 0
            for title, url in data:
                if url not in existing_urls:
                    ws.append([title, url])
                    new_data_count += 1

            wb.save(filename)
            print(f'新增 {new_data_count} 条数据，总数据量: {ws.max_row - 1} 条')
            return new_data_count

        else:
            # 新建文件或覆盖写入
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(['标题', '链接'])

            for title, url in data:
                ws.append([title, url])

            wb.save(filename)
            print(f'新建文件，写入 {len(data)} 条数据')
            return len(data)

    except Exception as e:
        print(f'写入Excel文件时出错: {e}')
        return 0


def scrape_data():
    """爬取数据的主函数"""
    all_data = []  # 临时存储当前批次的数据
    total_saved = 0  # 记录已保存的总数据量

    # 检查是否存在已有的Excel文件
    filename = '惠农网数据.xlsx'
    if os.path.exists(filename):
        wb = openpyxl.load_workbook(filename)
        ws = wb.active
        print(f'检测到已有数据文件，当前数据量: {ws.max_row - 1} 条')

    for i in range(1, 501):
        try:
            # 构建分页URL
            if i == 1:
                page_url = base_url
            else:
                page_url = f'https://www.cnhnb.com/p/sczw-0-0-0-0-{i}/'  # 根据实际分页格式调整

            print(f'正在爬取第{i}页...')

            response = requests.get(page_url, headers=headers)
            response.encoding = 'utf-8'  # 设置编码
            html_content = response.text

            # 解析HTML
            content = etree.HTML(html_content)

            # 提取商品编号
            urls_num = re.findall(r'href="/gongying/(\d+)/"', html_content)

            # 提取标题
            title_elements = content.xpath('//div[@class="shop-image"]//img/@title')
            titles = [title.strip() for title in title_elements if title.strip()]

            # 构建完整URL
            urls = [f'https://www.cnhnb.com/gongying/{num}/' for num in urls_num]

            # 确保标题和URL数量一致
            min_length = min(len(titles), len(urls))
            for j in range(min_length):
                all_data.append((titles[j], urls[j]))

            print(f'第{i}页爬取完成，获取到{min_length}条数据，当前批次累计{len(all_data)}条数据')

            # 每10页保存一次数据并清空列表
            if i % 10 == 0 and all_data:
                saved_count = write_to_excel(all_data, filename)
                total_saved += saved_count
                print(f'已保存前{i}页数据，本次保存{saved_count}条，累计保存{total_saved}条')
                all_data = []  # 清空列表，释放内存

            # 随机延迟
            sleep_time = random.randint(2, 4)
            time.sleep(sleep_time)

        except Exception as e:
            print(f'第{i}页爬取失败: {e}')
            continue

    # 处理剩余未保存的数据（如果总页数不是10的倍数）
    if all_data:
        saved_count = write_to_excel(all_data, filename)
        total_saved += saved_count
        print(f'保存剩余数据{len(all_data)}条，累计保存{total_saved}条')

    print('所有数据爬取完成！')


if __name__ == "__main__":
    scrape_data()