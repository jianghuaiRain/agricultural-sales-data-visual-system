import requests
from lxml import etree
import random
import re
import time
import openpyxl
import os

# 定义请求头
headers = {
    'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/142.0.0.0 Safari/537.36 Edg/142.0.0.0'}

# 定义所有品类的基础URL（根据网站实际品类补充，示例仅列部分）
category_urls = {
    "蔬菜": "https://www.cnhnb.com/p/sczw/",
    "禽畜肉蛋": "https://www.cnhnb.com/p/qcrd/",
    "农副加工": "https://www.cnhnb.com/p/nfjg/",
    "水产": "https://www.cnhnb.com/p/shuic/",
    "粮油米面": "https://www.cnhnb.com/p/lymm/",
    "水果":"https://www.cnhnb.com/p/sgzw/",
    "种子种苗":"https://www.cnhnb.com/p/zzzm/",

    # 可继续添加其他品类的基础URL...
}


def write_to_excel(data, filename='惠农网全品类数据.xlsx', is_append=True):
    """将数据写入Excel文件，支持增量写入，新增“品类”列"""
    try:
        if os.path.exists(filename) and is_append:
            wb = openpyxl.load_workbook(filename)
            ws = wb.active
            existing_urls = set()
            for row in range(2, ws.max_row + 1):
                url_cell = ws.cell(row=row, column=3).value  # 第3列是URL
                if url_cell:
                    existing_urls.add(url_cell)
            new_data_count = 0
            for category, title, url in data:
                if url not in existing_urls:
                    ws.append([category, title, url])
                    new_data_count += 1
            wb.save(filename)
            print(f'新增 {new_data_count} 条数据，总数据量: {ws.max_row - 1} 条')
            return new_data_count
        else:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(['品类', '标题', '链接'])  # 新增“品类”列
            for category, title, url in data:
                ws.append([category, title, url])
            wb.save(filename)
            print(f'新建文件，写入 {len(data)} 条数据')
            return len(data)
    except Exception as e:
        print(f'写入Excel文件时出错: {e}')
        return 0


def scrape_category(category, base_url):
    """爬取单个品类的所有分页数据"""
    all_data = []
    for i in range(1, 30):  # 假设最多爬500页，可根据实际情况调整
        try:
            if i == 1:
                page_url = base_url
            else:
                # 分页URL格式需与网站实际规则匹配，这里以“禽畜肉蛋”为例，其他品类可能需要调整
                page_url = f'{base_url[:-1]}-{i}/'  # 如禽畜肉蛋的分页是https://www.cnhnb.com/p/qcrd-2/

            print(f'正在爬取【{category}】第{i}页...')
            response = requests.get(page_url, headers=headers)
            response.encoding = 'utf-8'
            html_content = response.text

            # 提取商品编号和标题
            urls_num = re.findall(r'href="/gongying/(\d+)/"', html_content)
            title_elements = etree.HTML(html_content).xpath('//div[@class="shop-image"]//img/@title')
            titles = [title.strip() for title in title_elements if title.strip()]

            # 构建完整URL并匹配数据
            urls = [f'https://www.cnhnb.com/gongying/{num}/' for num in urls_num]
            min_length = min(len(titles), len(urls))
            for j in range(min_length):
                all_data.append((category, titles[j], urls[j]))

            print(f'【{category}】第{i}页爬取完成，当前批次累计{len(all_data)}条数据')

            # 每10页保存一次（可选，也可爬完品类再统一保存）
            if i % 10 == 0 and all_data:
                write_to_excel(all_data, filename='惠农网全品类数据.xlsx')
                all_data = []

            time.sleep(random.randint(2, 4))  # 随机延迟
        except Exception as e:
            print(f'【{category}】第{i}页爬取失败: {e}')
            break  # 某页失败时，跳出该品类的循环，避免无限尝试
    return all_data


def scrape_all_categories():
    """爬取所有品类的主函数"""
    total_saved = 0
    filename = '惠农网全品类数据.xlsx'
    for category, base_url in category_urls.items():
        print(f'开始爬取品类：{category}')
        category_data = scrape_category(category, base_url)
        if category_data:
            saved_count = write_to_excel(category_data, filename)
            total_saved += saved_count
            print(f'【{category}】爬取完成，累计保存{total_saved}条数据\n')
    print('所有品类爬取完成！')


if __name__ == "__main__":
    scrape_all_categories()